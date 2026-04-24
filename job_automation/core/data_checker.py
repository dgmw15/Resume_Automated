"""
core/data_checker.py — Deterministic data completeness audit and optional backfill.

Two execution modes (config-driven via data_checker.mode):
  audit_only — reads workbook, classifies gaps, emits completeness report. No writes.
  recover    — same checks, then backfills recoverable/derived fields, emits both reports.

Recovery waterfall (in recover mode):
  1. Local deterministic: reparse salary_raw with salary_parser; mirror min↔max.
  2. Portal refetch (Phase C): re-fetch URL when allow_portal_refetch=true in config.
  3. Unresolved: tag row with explicit reason string.

Recovery outcome states:
  COMPLETE          — all required fields valid, no action taken
  RECOVERED_LOCAL   — gap filled by local deterministic logic
  RECOVERED_REFETCH — gap filled by re-fetching source URL
  UNRESOLVED        — gap detected but could not be filled; reason recorded
  SKIPPED_NO_URL    — row has no URL; refetch structurally impossible
  ERROR_FETCH       — refetch attempted but network/selector call failed

Safety rules (non-negotiable):
  - dry_run=True by default; no writes until explicitly disabled
  - timestamped backup copy created before any mutation
  - only changed rows are written; unchanged rows are read-only
  - idempotent reruns: COMPLETE rows are never touched
  - row-level failures do not abort the run
"""
from __future__ import annotations

import asyncio
import json
import logging
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from core.salary_parser import parse_salary_range

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Recovery outcome states
# ---------------------------------------------------------------------------
COMPLETE = "COMPLETE"
RECOVERED_LOCAL = "RECOVERED_LOCAL"
RECOVERED_REFETCH = "RECOVERED_REFETCH"
UNRESOLVED = "UNRESOLVED"
SKIPPED_NO_URL = "SKIPPED_NO_URL"
ERROR_FETCH = "ERROR_FETCH"

# ---------------------------------------------------------------------------
# Field classifications
# ---------------------------------------------------------------------------
CRITICAL = "critical"
RECOVERABLE = "recoverable"
DERIVED = "derived"

# ---------------------------------------------------------------------------
# Missing-value gap types
# ---------------------------------------------------------------------------
TRULY_MISSING = "truly_missing"
SEMANTICALLY_MISSING = "semantically_missing"
INCONSISTENT = "inconsistent"

# Salary status values that signal semantic absence
_SALARY_BAD_STATUS: set[str] = {"MISSING", "AMBIGUOUS", "ERROR"}

# ---------------------------------------------------------------------------
# Field contracts
# ---------------------------------------------------------------------------

#: Trawl workbook (trawl_results.xlsx) — detected by presence of "portal" column.
TRAWL_CONTRACT: dict[str, str] = {
    "id": CRITICAL,
    "portal": CRITICAL,
    "role": CRITICAL,
    "url": CRITICAL,
    "company": RECOVERABLE,
    "raw_description": RECOVERABLE,
    "salary_raw": RECOVERABLE,
    "salary_min": DERIVED,
    "salary_max": DERIVED,
    "salary_currency": DERIVED,
    "salary_period": DERIVED,
    "salary_status": DERIVED,
}

#: Tracker workbook (Database.xlsx) — Phase D only; detected by "portal_name" column.
TRACKER_CONTRACT: dict[str, str] = {
    "id": CRITICAL,
    "portal_name": CRITICAL,
    "role": CRITICAL,
    "url": CRITICAL,
    "company": RECOVERABLE,
    "raw_description": RECOVERABLE,
    "status": CRITICAL,
    "salary_raw": RECOVERABLE,
    "salary_min": DERIVED,
    "salary_max": DERIVED,
    "salary_currency": DERIVED,
    "salary_period": DERIVED,
    "salary_status": DERIVED,
}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class FieldGap:
    field_name: str
    classification: str  # CRITICAL | RECOVERABLE | DERIVED
    gap_type: str        # TRULY_MISSING | SEMANTICALLY_MISSING | INCONSISTENT
    detail: str = ""


@dataclass
class RowResult:
    row_id: str
    outcome: str                                        # one of the six outcome constants
    gaps: list[FieldGap] = field(default_factory=list)
    recovery_actions: list[str] = field(default_factory=list)
    reason: str = ""                                    # required for non-COMPLETE outcomes


@dataclass
class WorkbookReport:
    workbook_path: str
    total_rows: int
    mode: str
    dry_run: bool
    results: list[RowResult] = field(default_factory=list)

    # ------------------------------------------------------------------
    # Computed summaries
    # ------------------------------------------------------------------

    @property
    def outcome_counts(self) -> dict[str, int]:
        counts: dict[str, int] = {
            COMPLETE: 0, RECOVERED_LOCAL: 0, RECOVERED_REFETCH: 0,
            UNRESOLVED: 0, SKIPPED_NO_URL: 0, ERROR_FETCH: 0,
        }
        for r in self.results:
            counts[r.outcome] = counts.get(r.outcome, 0) + 1
        return counts

    @property
    def field_missing_pct(self) -> dict[str, float]:
        """Percentage of rows where each field has any kind of gap."""
        if not self.total_rows:
            return {}
        counts: dict[str, int] = {}
        for r in self.results:
            seen: set[str] = set()
            for g in r.gaps:
                if g.field_name not in seen:
                    counts[g.field_name] = counts.get(g.field_name, 0) + 1
                    seen.add(g.field_name)
        return {f: round(c / self.total_rows * 100, 2) for f, c in counts.items()}

    @property
    def portal_breakdown(self) -> dict[str, dict[str, int]]:
        """Outcome counts grouped by portal (uses portal/portal_name field if present)."""
        breakdown: dict[str, dict[str, int]] = {}
        # We don't store raw row values in RowResult; breakdown needs the caller to supply it.
        # This property is populated externally by DataChecker._check_workbook.
        return breakdown

    def to_dict(self) -> dict:
        return {
            "workbook": self.workbook_path,
            "generated_at": datetime.utcnow().isoformat(),
            "mode": self.mode,
            "dry_run": self.dry_run,
            "total_rows": self.total_rows,
            "outcome_counts": self.outcome_counts,
            "field_missing_pct": self.field_missing_pct,
            "portal_breakdown": self.portal_breakdown,
            "unresolved_details": [
                {
                    "row_id": r.row_id,
                    "outcome": r.outcome,
                    "reason": r.reason,
                    "gaps": [
                        {"field": g.field_name, "type": g.gap_type, "detail": g.detail}
                        for g in r.gaps
                    ],
                }
                for r in self.results
                if r.outcome in (UNRESOLVED, SKIPPED_NO_URL, ERROR_FETCH)
            ],
        }

    def to_recovery_dict(self) -> dict:
        recovered = [r for r in self.results if r.outcome in (RECOVERED_LOCAL, RECOVERED_REFETCH)]
        unresolved = [r for r in self.results if r.outcome in (UNRESOLVED, SKIPPED_NO_URL, ERROR_FETCH)]
        return {
            "workbook": self.workbook_path,
            "generated_at": datetime.utcnow().isoformat(),
            "mode": self.mode,
            "dry_run": self.dry_run,
            "rows_attempted": len(self.results) - self.outcome_counts[COMPLETE],
            "rows_fixed": len(recovered),
            "rows_unresolved": len(unresolved),
            "recovery_breakdown": self.outcome_counts,
            "unresolved_by_reason": _group_by_reason(unresolved),
            "recovery_actions": [
                {"row_id": r.row_id, "outcome": r.outcome, "actions": r.recovery_actions}
                for r in recovered
            ],
        }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_absent(val: Any) -> bool:
    """Return True if the cell value counts as truly missing."""
    if val is None:
        return True
    if isinstance(val, str) and not val.strip():
        return True
    return False


def _summarise_gaps(gaps: list[FieldGap]) -> str:
    if not gaps:
        return ""
    return "; ".join(f"{g.field_name}({g.gap_type})" for g in gaps)


def _changed_cells(original: dict, updated: dict) -> dict:
    """Return only the key/value pairs that differ from original."""
    return {k: v for k, v in updated.items() if v != original.get(k)}


def _group_by_reason(results: list[RowResult]) -> dict[str, int]:
    reasons: dict[str, int] = {}
    for r in results:
        key = r.reason or "unknown"
        reasons[key] = reasons.get(key, 0) + 1
    return reasons


def _make_backup(path: Path) -> Path:
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    backup = path.parent / f"{path.stem}_backup_{ts}{path.suffix}"
    shutil.copy2(path, backup)
    logger.info("[data_checker] Backup created: %s", backup)
    return backup


# ---------------------------------------------------------------------------
# DataChecker
# ---------------------------------------------------------------------------

class DataChecker:
    """
    Deterministic workbook completeness checker.

    Usage (from Python):
        checker = DataChecker.from_config(config)
        reports = asyncio.run(checker.run())

    Usage (standalone):
        python check_data.py [--mode recover] [--no-dry-run]
    """

    def __init__(self, config: dict) -> None:
        cfg = config.get("data_checker", {})
        sal = config.get("salary", {})

        self._enabled: bool = bool(cfg.get("enabled", False))
        self._dry_run: bool = bool(cfg.get("dry_run", True))
        self._mode: str = str(cfg.get("mode", "audit_only"))
        self._targets: list[str] = list(cfg.get("targets", []))
        self._write_backup: bool = bool(cfg.get("write_backup", True))
        self._backup_days: int = int(cfg.get("backup_retention_days", 7))

        backfill = cfg.get("backfill", {})
        self._allow_refetch: bool = bool(backfill.get("allow_portal_refetch", False))
        self._refetch_portals: list[str] = [
            p.lower() for p in backfill.get("portals_allowed_for_refetch", ["careersfuture"])
        ]
        self._max_rows: int = int(backfill.get("max_rows_per_run", 50))
        self._reason_required: bool = bool(backfill.get("unresolved_reason_required", True))

        self._report_tpl: str = str(
            cfg.get("report_path", "output/logs/completeness_report_{date}.json")
        )
        self._recovery_tpl: str = str(
            cfg.get("recovery_report_path", "output/logs/recovery_report_{date}.json")
        )
        self._issue_wb: str = str(cfg.get("issue_workbook_path", ""))

        self._default_currency: str = str(sal.get("default_currency", "SGD"))
        self._period_inference: bool = bool(sal.get("enable_period_inference", True))

    # ------------------------------------------------------------------
    # Factory
    # ------------------------------------------------------------------

    @classmethod
    def from_config(cls, config: dict) -> "DataChecker":
        return cls(config)

    # ------------------------------------------------------------------
    # Public entry points
    # ------------------------------------------------------------------

    def run_sync(self) -> list[WorkbookReport]:
        """
        Synchronous entry point for standalone CLI and non-async callers.
        Uses asyncio.run() internally; do not call from inside an async context.
        """
        if not self._enabled:
            logger.info("[data_checker] Disabled in config — skipping.")
            return []
        return asyncio.run(self.run())

    async def run(self) -> list[WorkbookReport]:
        """
        Full async run — supports both local and portal-refetch recovery.
        Returns one WorkbookReport per target workbook processed.
        """
        if not self._enabled:
            logger.info("[data_checker] Disabled in config — skipping.")
            return []

        date_str = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        reports: list[WorkbookReport] = []

        for target in self._targets:
            path = Path(target)
            if not path.exists():
                logger.warning("[data_checker] Target not found: %s — skipping.", path)
                continue
            try:
                report = await self._check_workbook(path)
                reports.append(report)
            except Exception as exc:
                logger.error("[data_checker] Failed to process %s: %s", path, exc, exc_info=True)

        self._write_reports(reports, date_str)
        return reports

    # ------------------------------------------------------------------
    # Workbook processing
    # ------------------------------------------------------------------

    async def _check_workbook(self, path: Path) -> WorkbookReport:
        wb = load_workbook(path, data_only=True)

        # Detect sheet name
        sheet_name = "Trawl" if "Trawl" in wb.sheetnames else "Jobs"
        ws = wb[sheet_name]

        # Build 0-based column index {col_name: index}
        header: list[Any] = [
            ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
        ]
        col_idx: dict[str, int] = {
            str(name): i for i, name in enumerate(header) if name is not None
        }

        # Select field contract based on which portal column is present
        contract = TRAWL_CONTRACT if "portal" in col_idx else TRACKER_CONTRACT
        portal_field = "portal" if "portal" in col_idx else "portal_name"

        total_rows = max(ws.max_row - 1, 0)
        report = WorkbookReport(
            workbook_path=str(path),
            total_rows=total_rows,
            mode=self._mode,
            dry_run=self._dry_run,
        )

        # Track per-portal missing counts for breakdown
        portal_counts: dict[str, dict[str, int]] = {}

        # Rows that were locally recovered and need writing back
        rows_to_write: dict[int, dict[str, Any]] = {}

        refetch_count = 0

        for row_excel in range(2, ws.max_row + 1):
            row_vals = [
                ws.cell(row=row_excel, column=c).value
                for c in range(1, ws.max_column + 1)
            ]
            row_dict: dict[str, Any] = {
                str(header[i]): row_vals[i] for i in range(len(header)) if header[i] is not None
            }

            row_id = str(row_dict.get("id") or f"row_{row_excel}")
            portal_name = str(row_dict.get(portal_field) or "unknown").lower()

            # ----------------------------------------------------------
            # Classify gaps
            # ----------------------------------------------------------
            gaps = self._classify_gaps(row_dict, contract)

            # Update portal breakdown
            outcome_key = COMPLETE if not gaps else UNRESOLVED
            if portal_name not in portal_counts:
                portal_counts[portal_name] = {k: 0 for k in (COMPLETE, UNRESOLVED)}
            portal_counts[portal_name][outcome_key] = (
                portal_counts[portal_name].get(outcome_key, 0) + 1
            )

            if not gaps:
                report.results.append(RowResult(row_id=row_id, outcome=COMPLETE))
                continue

            # ----------------------------------------------------------
            # Audit-only mode: record without recovery
            # ----------------------------------------------------------
            if self._mode == "audit_only":
                report.results.append(RowResult(
                    row_id=row_id,
                    outcome=UNRESOLVED,
                    gaps=gaps,
                    reason=_summarise_gaps(gaps),
                ))
                continue

            # ----------------------------------------------------------
            # Recover mode: attempt local recovery first
            # ----------------------------------------------------------
            updated, actions = self._try_local_recovery(row_dict, gaps, contract)
            remaining = self._classify_gaps(updated, contract)

            if not remaining:
                # Local recovery resolved all gaps
                if not self._dry_run:
                    changed = _changed_cells(row_dict, updated)
                    if changed:
                        rows_to_write[row_excel] = changed
                report.results.append(RowResult(
                    row_id=row_id,
                    outcome=RECOVERED_LOCAL,
                    gaps=gaps,
                    recovery_actions=actions,
                ))
                continue

            # ----------------------------------------------------------
            # Portal refetch (Phase C) — only if still unresolved
            # ----------------------------------------------------------
            if self._allow_refetch and refetch_count < self._max_rows:
                url = str(updated.get("url") or "").strip()
                portal = str(updated.get(portal_field) or "").lower()

                if not url:
                    report.results.append(RowResult(
                        row_id=row_id,
                        outcome=SKIPPED_NO_URL,
                        gaps=remaining,
                        recovery_actions=actions,
                        reason="Row has no URL; refetch not possible.",
                    ))
                    continue

                if portal not in self._refetch_portals:
                    report.results.append(RowResult(
                        row_id=row_id,
                        outcome=UNRESOLVED,
                        gaps=remaining,
                        recovery_actions=actions,
                        reason=f"Portal '{portal}' not in refetch allowlist.",
                    ))
                    continue

                refetch_count += 1
                try:
                    refetched, ref_actions = await self._try_refetch(updated, remaining, portal)
                    still_remaining = self._classify_gaps(refetched, contract)
                    if not still_remaining:
                        if not self._dry_run:
                            changed = _changed_cells(row_dict, refetched)
                            if changed:
                                rows_to_write[row_excel] = changed
                        report.results.append(RowResult(
                            row_id=row_id,
                            outcome=RECOVERED_REFETCH,
                            gaps=gaps,
                            recovery_actions=actions + ref_actions,
                        ))
                    else:
                        report.results.append(RowResult(
                            row_id=row_id,
                            outcome=UNRESOLVED,
                            gaps=still_remaining,
                            recovery_actions=actions + ref_actions,
                            reason=_summarise_gaps(still_remaining),
                        ))
                except Exception as exc:
                    report.results.append(RowResult(
                        row_id=row_id,
                        outcome=ERROR_FETCH,
                        gaps=remaining,
                        recovery_actions=actions,
                        reason=f"Refetch failed: {exc}",
                    ))
                continue

            # ----------------------------------------------------------
            # Cannot recover
            # ----------------------------------------------------------
            report.results.append(RowResult(
                row_id=row_id,
                outcome=UNRESOLVED,
                gaps=remaining,
                recovery_actions=actions,
                reason=_summarise_gaps(remaining),
            ))

        # ------------------------------------------------------------------
        # Write back changed rows
        # ------------------------------------------------------------------
        if rows_to_write:
            if self._write_backup:
                _make_backup(path)
            for row_excel, changes in rows_to_write.items():
                for col_name, new_val in changes.items():
                    if col_name in col_idx:
                        ws.cell(row=row_excel, column=col_idx[col_name] + 1, value=new_val)
            wb.save(path)
            logger.info(
                "[data_checker] %s — wrote %d changed rows.",
                path.name, len(rows_to_write),
            )
        elif self._dry_run and self._mode == "recover":
            recoverable = sum(
                1 for r in report.results
                if r.outcome in (RECOVERED_LOCAL, RECOVERED_REFETCH)
            )
            if recoverable:
                logger.info(
                    "[data_checker] DRY RUN — %d rows would be updated in %s (no writes).",
                    recoverable, path.name,
                )

        # Attach portal breakdown to report
        report.portal_breakdown.update(portal_counts)  # type: ignore[attr-defined]

        return report

    # ------------------------------------------------------------------
    # Gap classification
    # ------------------------------------------------------------------

    def _classify_gaps(
        self, row: dict, contract: dict[str, str]
    ) -> list[FieldGap]:
        gaps: list[FieldGap] = []

        for field_name, classification in contract.items():
            val = row.get(field_name)

            # 1. Truly missing — null or empty string
            if _is_absent(val):
                gaps.append(FieldGap(
                    field_name=field_name,
                    classification=classification,
                    gap_type=TRULY_MISSING,
                    detail=f"'{field_name}' is null or empty",
                ))
                continue  # Don't add semantic check on top of truly missing

            # 2. Semantically missing — salary_status signals parse failure
            if field_name == "salary_status" and str(val) in _SALARY_BAD_STATUS:
                gaps.append(FieldGap(
                    field_name=field_name,
                    classification=classification,
                    gap_type=SEMANTICALLY_MISSING,
                    detail=f"salary_status={val!r} indicates unparsed salary",
                ))

        # 3. Inconsistencies — checked only when individual fields are present
        gaps.extend(self._check_inconsistencies(row, contract))
        return gaps

    @staticmethod
    def _check_inconsistencies(
        row: dict, contract: dict[str, str]
    ) -> list[FieldGap]:
        """Detect values that contradict each other."""
        issues: list[FieldGap] = []
        salary_fields = {"salary_min", "salary_max", "salary_currency", "salary_status"}
        if not salary_fields.issubset(contract):
            return issues

        sal_min = row.get("salary_min")
        sal_max = row.get("salary_max")
        sal_currency = row.get("salary_currency")
        sal_status = row.get("salary_status")

        # salary_min > salary_max
        if not _is_absent(sal_min) and not _is_absent(sal_max):
            try:
                if float(sal_min) > float(sal_max):
                    issues.append(FieldGap(
                        field_name="salary_min",
                        classification=DERIVED,
                        gap_type=INCONSISTENT,
                        detail=f"salary_min ({sal_min}) > salary_max ({sal_max})",
                    ))
            except (ValueError, TypeError):
                pass

        # currency is null while salary_min has a value
        if not _is_absent(sal_min) and _is_absent(sal_currency):
            issues.append(FieldGap(
                field_name="salary_currency",
                classification=DERIVED,
                gap_type=INCONSISTENT,
                detail="salary_currency is null but salary_min has a value",
            ))

        # salary_status=OK but both numeric fields are null
        if str(sal_status or "") == "OK" and _is_absent(sal_min) and _is_absent(sal_max):
            issues.append(FieldGap(
                field_name="salary_status",
                classification=DERIVED,
                gap_type=INCONSISTENT,
                detail="salary_status=OK but both salary_min and salary_max are null",
            ))

        return issues

    # ------------------------------------------------------------------
    # Local deterministic recovery
    # ------------------------------------------------------------------

    def _try_local_recovery(
        self,
        row: dict,
        gaps: list[FieldGap],
        contract: dict[str, str],
    ) -> tuple[dict, list[str]]:
        """
        Attempt to fill salary derived fields without any network call.

        Returns (updated_row, actions_taken).
        """
        updated = dict(row)
        actions: list[str] = []

        gap_fields = {g.field_name for g in gaps}
        salary_derived = {"salary_min", "salary_max", "salary_currency", "salary_period", "salary_status"}
        has_salary_contract = salary_derived.issubset(contract)

        if not has_salary_contract:
            return updated, actions

        # Recovery 1: reparse salary_raw if derived salary fields are gapped
        if gap_fields & salary_derived:
            raw = str(updated.get("salary_raw") or "").strip()
            if raw:
                try:
                    result = parse_salary_range(
                        raw_text=raw,
                        default_currency=self._default_currency,
                        enable_period_inference=self._period_inference,
                    )
                    if result.salary_status == "OK":
                        updated["salary_min"] = result.salary_min
                        updated["salary_max"] = result.salary_max
                        updated["salary_currency"] = result.salary_currency
                        updated["salary_period"] = result.salary_period
                        updated["salary_status"] = result.salary_status
                        actions.append(
                            f"Reparsed salary_raw={raw!r}: "
                            f"min={result.salary_min}, max={result.salary_max}, "
                            f"status={result.salary_status}"
                        )
                except Exception as exc:
                    logger.debug(
                        "[data_checker] salary_raw reparse failed for raw=%r: %s", raw, exc
                    )

        # Recovery 2: mirror min ↔ max when only one side is present
        sal_min = updated.get("salary_min")
        sal_max = updated.get("salary_max")
        if not _is_absent(sal_min) and _is_absent(sal_max):
            updated["salary_max"] = sal_min
            actions.append(f"Mirrored salary_max = salary_min = {sal_min}")
        elif not _is_absent(sal_max) and _is_absent(sal_min):
            updated["salary_min"] = sal_max
            actions.append(f"Mirrored salary_min = salary_max = {sal_max}")

        return updated, actions

    # ------------------------------------------------------------------
    # Portal refetch (Phase C — CareersFuture supported)
    # ------------------------------------------------------------------

    async def _try_refetch(
        self,
        row: dict,
        gaps: list[FieldGap],
        portal: str,
    ) -> tuple[dict, list[str]]:
        """
        Re-fetch salary data from the source URL for a given portal.

        Currently supports: careersfuture.
        Raises RuntimeError on failure so the caller can tag the row ERROR_FETCH.
        """
        url = str(row.get("url") or "").strip()
        updated = dict(row)
        actions: list[str] = []

        if portal == "careersfuture":
            from playwright.async_api import async_playwright
            from adapters.careersfuture import CSS_SALARY_RANGE, CSS_SALARY_MIN, CSS_SALARY_MAX

            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                try:
                    page = await browser.new_page()
                    await page.goto(url, timeout=20_000)
                    await page.wait_for_load_state("domcontentloaded")

                    raw = ""
                    min_text = ""
                    max_text = ""

                    try:
                        raw = (await page.locator(CSS_SALARY_RANGE).text_content(timeout=5_000)) or ""
                    except Exception:
                        pass
                    try:
                        min_text = (await page.locator(CSS_SALARY_MIN).nth(0).text_content(timeout=3_000)) or ""
                    except Exception:
                        pass
                    try:
                        max_text = (await page.locator(CSS_SALARY_MAX).nth(1).text_content(timeout=3_000)) or ""
                    except Exception:
                        pass

                    result = parse_salary_range(
                        raw_text=raw,
                        min_text=min_text,
                        max_text=max_text,
                        default_currency=self._default_currency,
                        enable_period_inference=self._period_inference,
                    )
                    updated["salary_raw"] = result.salary_raw
                    updated["salary_min"] = result.salary_min
                    updated["salary_max"] = result.salary_max
                    updated["salary_currency"] = result.salary_currency
                    updated["salary_period"] = result.salary_period
                    updated["salary_status"] = result.salary_status
                    actions.append(
                        f"Refetched salary from {url}: status={result.salary_status}"
                    )
                except Exception as exc:
                    raise RuntimeError(f"CareersFuture refetch error: {exc}") from exc
                finally:
                    await browser.close()
        else:
            raise NotImplementedError(f"Refetch not implemented for portal: {portal!r}")

        return updated, actions

    # ------------------------------------------------------------------
    # Report writers
    # ------------------------------------------------------------------

    def _write_reports(self, reports: list[WorkbookReport], date_str: str) -> None:
        for report in reports:
            self._write_json(report.to_dict(), self._report_tpl, date_str)
            self._write_json(report.to_recovery_dict(), self._recovery_tpl, date_str)
            if self._issue_wb:
                self._write_issue_workbook(report, date_str)

        for report in reports:
            counts = report.outcome_counts
            logger.info(
                "[data_checker] %s — rows=%d  COMPLETE=%d  RECOVERED_LOCAL=%d  "
                "RECOVERED_REFETCH=%d  UNRESOLVED=%d  SKIPPED_NO_URL=%d  ERROR_FETCH=%d",
                Path(report.workbook_path).name,
                report.total_rows,
                counts[COMPLETE],
                counts[RECOVERED_LOCAL],
                counts[RECOVERED_REFETCH],
                counts[UNRESOLVED],
                counts[SKIPPED_NO_URL],
                counts[ERROR_FETCH],
            )

    @staticmethod
    def _write_json(data: dict, path_tpl: str, date_str: str) -> None:
        path = Path(path_tpl.replace("{date}", date_str))
        path.parent.mkdir(parents=True, exist_ok=True)
        try:
            path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
            logger.info("[data_checker] Report written: %s", path)
        except Exception as exc:
            logger.error("[data_checker] Could not write report %s: %s", path, exc)

    def _write_issue_workbook(self, report: WorkbookReport, date_str: str) -> None:
        from openpyxl import Workbook as OXLWorkbook
        from openpyxl.styles import Font, PatternFill

        non_complete = [r for r in report.results if r.outcome != COMPLETE]
        if not non_complete:
            return

        path = Path(self._issue_wb.replace("{date}", date_str))
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = OXLWorkbook()
        ws = wb.active
        ws.title = "Unresolved"

        headers = ["row_id", "outcome", "reason", "gap_fields", "gap_types"]
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = fill
            cell.font = font

        for row_num, result in enumerate(non_complete, 2):
            ws.cell(row=row_num, column=1, value=result.row_id)
            ws.cell(row=row_num, column=2, value=result.outcome)
            ws.cell(row=row_num, column=3, value=result.reason)
            ws.cell(row=row_num, column=4, value=", ".join(g.field_name for g in result.gaps))
            ws.cell(row=row_num, column=5, value=", ".join(g.gap_type for g in result.gaps))

        wb.save(path)
        logger.info("[data_checker] Issue workbook written: %s", path)

    # ------------------------------------------------------------------
    # Backup retention cleanup
    # ------------------------------------------------------------------

    def cleanup_old_backups(self, workbook_path: Path) -> None:
        """Delete backup copies older than backup_retention_days."""
        from datetime import timedelta

        cutoff = datetime.utcnow() - timedelta(days=self._backup_days)
        parent = workbook_path.parent
        stem = workbook_path.stem
        suffix = workbook_path.suffix
        pattern = f"{stem}_backup_*{suffix}"

        for backup in parent.glob(pattern):
            try:
                mtime = datetime.utcfromtimestamp(backup.stat().st_mtime)
                if mtime < cutoff:
                    backup.unlink()
                    logger.info("[data_checker] Deleted old backup: %s", backup)
            except Exception as exc:
                logger.warning("[data_checker] Could not delete backup %s: %s", backup, exc)
