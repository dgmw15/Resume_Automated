"""
trawl.py — Standalone scraping script for testing and data collection.

Scrapes job listings from enabled portals and logs everything to an Excel file.
The output Excel can be fed directly into the AI pipeline later.

Usage:
    python trawl.py                          # use defaults from config.yaml
    python trawl.py --pages 3                # scrape 3 pages per role
    python trawl.py --portal careersfuture   # only one portal
    python trawl.py --output my_jobs.xlsx    # custom output file
    python trawl.py --no-descriptions        # skip fetching full JD text (faster)

Output columns:
    id, scraped_at, portal, role, company, url, page_num,
    raw_description, description_status, notes
"""
from __future__ import annotations

import argparse
import asyncio
import logging
import random
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from playwright.async_api import async_playwright
from core.login_utils import wait_with_progress
from core.salary_parser import parse_salary_range

# ---------------------------------------------------------------------------
# Browser check — give a plain-English fix if Chromium isn't installed
# ---------------------------------------------------------------------------
_CHROMIUM_PATHS = [
    Path.home() / "AppData/Local/ms-playwright",   # Windows
    Path.home() / ".cache/ms-playwright",           # Linux/Mac
]

def _check_browser_installed() -> None:
    """Exit with a clear message if Playwright's Chromium hasn't been downloaded."""
    found = any(p.exists() and any(p.iterdir()) for p in _CHROMIUM_PATHS if p.exists())
    if not found:
        print()
        print("=" * 60)
        print("  SETUP NEEDED — browser not installed")
        print("=" * 60)
        print()
        print("  The script uses a built-in browser to open job sites,")
        print("  but the browser files haven't been downloaded yet.")
        print()
        print("  Fix: run this command once in your terminal:")
        print()
        print("    .venv\\Scripts\\playwright.exe install chromium")
        print()
        print("  Then re-run trawl.py.")
        print()
        sys.exit(1)

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("trawl.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CONFIG_PATH   = Path("config.yaml")
ROLES_PATH    = Path("job_roles.xlsx")
SESSION_DIR   = Path(".sessions")
DEFAULT_OUT   = Path("trawl_results.xlsx")

COLUMNS = [
    "id", "scraped_at", "portal", "role", "company",
    "url", "page_num", "raw_description", "description_status", "notes", "skills", "continue",
    # Salary fields (S2)
    "salary_raw", "salary_min", "salary_max", "salary_currency", "salary_period", "salary_status",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def init_excel(path: Path) -> None:
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Trawl"
        for i, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        wb.save(path)
        logger.info("Created output file: %s", path)
        return

    wb = load_workbook(path)
    ws = wb.active
    existing = {ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)}
    for col in COLUMNS:
        if col not in existing:
            next_col = ws.max_column + 1
            cell = ws.cell(row=1, column=next_col, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
    wb.save(path)


def load_existing_urls(path: Path) -> set[str]:
    """Return all URLs already logged so we don't duplicate rows."""
    if not path.exists():
        return set()
    wb = load_workbook(path)
    ws = wb.active
    header = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}
    url_col = header.get("url")
    if url_col is None:
        return set()
    return {
        ws.cell(row=r, column=url_col).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=url_col).value
    }


def append_row(path: Path, row_data: dict) -> None:
    wb = load_workbook(path)
    ws = wb.active
    header = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}
    next_row = ws.max_row + 1
    for col_name, col_idx in header.items():
        value = row_data.get(col_name, "")
        if isinstance(value, datetime):
            value = value.isoformat()
        ws.cell(row=next_row, column=col_idx, value=value)
    wb.save(path)


# ---------------------------------------------------------------------------
# Role loader
# ---------------------------------------------------------------------------

def load_roles(config: dict) -> list[str]:
    roles_cfg = config.get("search", {}).get("roles_excel", {})
    excel_path = Path(roles_cfg.get("path", "job_roles.xlsx"))
    sheet  = roles_cfg.get("sheet", "Roles")
    column = roles_cfg.get("column", "Job Role")

    if excel_path.exists():
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet)
            roles = df[column].dropna().str.strip().tolist()
            if roles:
                logger.info("Loaded %d roles from %s", len(roles), excel_path)
                return roles
        except Exception as exc:
            logger.warning("Could not read %s: %s", excel_path, exc)

    fallback = config.get("search", {}).get("keywords", [])
    logger.info("Using %d fallback keywords from config.yaml", len(fallback))
    return fallback


# ---------------------------------------------------------------------------
# Portal scrapers
# ---------------------------------------------------------------------------

async def scrape_careersfuture(
    playwright,
    roles: list[str],
    pages_per_role: int,
    fetch_descriptions: bool,
    out_path: Path,
    existing_urls: set[str],
    delay_range: tuple[float, float],
) -> int:
    """Scrape CareersFuture and append rows to out_path. Returns total rows added."""
    from adapters.careersfuture import (
        CareersFutureAdapter,
        BASE_SEARCH_URL,
        CSS_JOB_CARD,
        CSS_CARD_TITLE,
        CSS_CARD_COMPANY,
        CSS_JOB_DESCRIPTION,
        CSS_SALARY_RANGE,
        CSS_SALARY_MIN,
        CSS_SALARY_MAX,
        MCF_HOME_URL,
        CSS_LOGIN_BTN,
        TIMEOUT_MS,
        LOGIN_TIMEOUT_MS,
    )

    session_dir = SESSION_DIR / "careersfuture"
    session_dir.mkdir(parents=True, exist_ok=True)

    try:
        context = await playwright.chromium.launch_persistent_context(
            user_data_dir=str(session_dir),
            headless=False,
            viewport={"width": 1280, "height": 800},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
        )
    except Exception as exc:
        if "Executable doesn't exist" in str(exc) or "spawn UNKNOWN" in str(exc):
            logger.error(
                "\n"
                "========================================\n"
                "  BROWSER NOT FOUND\n"
                "========================================\n"
                "  The built-in browser needs to be installed once.\n"
                "  Open a terminal in this folder and run:\n"
                "\n"
                "    .venv\\Scripts\\playwright.exe install chromium\n"
                "\n"
                "  Then re-run trawl.py.\n"
                "========================================"
            )
        else:
            logger.error("Failed to open browser: %s", exc)
        return 0

    added = 0
    try:
        # ── Login check ────────────────────────────────────────────────────
        login_page = await context.new_page()
        try:
            await login_page.goto(MCF_HOME_URL, timeout=TIMEOUT_MS)
        except Exception as exc:
            logger.error(
                "Could not open MyCareersFuture — check your internet connection. (%s)", exc
            )
            await login_page.close()
            return 0

        # Do not verify login state; just provide a manual 2-minute auth window.
        await login_page.wait_for_load_state("domcontentloaded")
        logger.info(
            "[careersfuture] Manual login window opened.\n"
            "  → If prompted, click 'Log In' and complete SingPass.\n"
            "  → Waiting up to 2 minutes before scraping starts."
        )
        try:
            await login_page.click(CSS_LOGIN_BTN, timeout=5_000)
            logger.info("[careersfuture] Login button clicked.")
        except Exception:
            logger.info("[careersfuture] Login button not clickable now; continue with manual flow.")

        await wait_with_progress(
            page=login_page,
            total_ms=LOGIN_TIMEOUT_MS,
            label="[careersfuture] Login time remaining",
        )
        logger.info("[careersfuture] Manual login window ended — starting scrape.")

        await login_page.close()

        # ── Scrape ─────────────────────────────────────────────────────────
        for role in roles:
            for page_num in range(1, pages_per_role + 1):
                url = BASE_SEARCH_URL.format(role=role, page=page_num)
                logger.info("[careersfuture] Role='%s' page=%d → %s", role, page_num, url)

                search_page = await context.new_page()
                try:
                    await search_page.goto(url, timeout=TIMEOUT_MS)
                    await search_page.wait_for_load_state("networkidle")

                    cards = await search_page.query_selector_all(CSS_JOB_CARD)
                    if not cards:
                        logger.info("[careersfuture] No cards on page %d — stopping role.", page_num)
                        await search_page.close()
                        break

                    logger.info("[careersfuture] Found %d cards on page %d.", len(cards), page_num)

                    listings = []
                    for card in cards:
                        try:
                            title_el   = await card.query_selector(CSS_CARD_TITLE)
                            company_el = await card.query_selector(CSS_CARD_COMPANY)
                            title   = (await title_el.inner_text()).strip()   if title_el   else "Unknown"
                            company = (await company_el.inner_text()).strip() if company_el else "Unknown"
                            href    = await card.get_attribute("href")
                            if href and not href.startswith("http"):
                                href = "https://www.mycareersfuture.gov.sg" + href

                            # Salary extraction from listing card (S3)
                            salary_raw = salary_min_text = salary_max_text = ""
                            try:
                                sal_el  = await card.query_selector(CSS_SALARY_RANGE)
                                min_el  = await card.query_selector(CSS_SALARY_MIN)
                                max_el  = await card.query_selector(CSS_SALARY_MAX)
                                if sal_el:
                                    salary_raw = (await sal_el.inner_text()).strip()
                                if min_el:
                                    salary_min_text = (await min_el.inner_text()).strip()
                                if max_el:
                                    salary_max_text = (await max_el.inner_text()).strip()
                            except Exception:
                                pass  # salary not found on card — fallback handled below

                            if href:
                                listings.append({
                                    "role": title, "company": company, "url": href,
                                    "salary_raw": salary_raw,
                                    "salary_min_text": salary_min_text,
                                    "salary_max_text": salary_max_text,
                                })
                        except Exception as exc:
                            logger.warning("[careersfuture] Card parse error: %s", exc)

                    await search_page.close()
                    _delay(delay_range)

                    # ── Fetch descriptions ──────────────────────────────
                    for listing in listings:
                        job_url = listing["url"]
                        if job_url in existing_urls:
                            logger.debug("[careersfuture] Skipping duplicate: %s", job_url)
                            continue

                        # Parse salary from listing card data
                        sal_cfg = config.get("salary", {})
                        sal_result = parse_salary_range(
                            raw_text=listing.get("salary_raw", ""),
                            min_text=listing.get("salary_min_text", ""),
                            max_text=listing.get("salary_max_text", ""),
                            default_currency=sal_cfg.get("default_currency", "SGD"),
                            enable_period_inference=sal_cfg.get("enable_period_inference", True),
                        )

                        row = {
                            "id":          str(uuid.uuid4()),
                            "scraped_at":  datetime.utcnow().isoformat(),
                            "portal":      "careersfuture",
                            "role":        listing["role"],
                            "company":     listing["company"],
                            "url":         job_url,
                            "page_num":    page_num,
                            "raw_description":   "",
                            "description_status": "PENDING",
                            "notes":       "",
                            # Salary (from card)
                            "salary_raw":      sal_result.salary_raw,
                            "salary_min":      sal_result.salary_min,
                            "salary_max":      sal_result.salary_max,
                            "salary_currency": sal_result.salary_currency,
                            "salary_period":   sal_result.salary_period,
                            "salary_status":   sal_result.salary_status,
                        }

                        if fetch_descriptions:
                            desc_page = await context.new_page()
                            try:
                                await desc_page.goto(job_url, timeout=TIMEOUT_MS)
                                await desc_page.wait_for_load_state("networkidle")
                                desc_el = await desc_page.query_selector(CSS_JOB_DESCRIPTION)
                                if desc_el:
                                    row["raw_description"]   = (await desc_el.inner_text()).strip()
                                    row["description_status"] = "OK"
                                else:
                                    row["description_status"] = "MISSING"
                                    row["notes"] = "Description element not found"

                                # Detail-page salary fallback (S3): only if card salary was missing
                                if sal_result.salary_status == "MISSING" and sal_cfg.get("capture_on_detail_fallback", True):
                                    try:
                                        sal_range_el = await desc_page.query_selector(CSS_SALARY_RANGE)
                                        sal_min_el   = await desc_page.query_selector(CSS_SALARY_MIN)
                                        sal_max_el   = await desc_page.query_selector(CSS_SALARY_MAX)
                                        raw_fb   = (await sal_range_el.inner_text()).strip() if sal_range_el else ""
                                        min_fb   = (await sal_min_el.inner_text()).strip()   if sal_min_el   else ""
                                        max_fb   = (await sal_max_el.inner_text()).strip()   if sal_max_el   else ""
                                        if raw_fb:
                                            fb_result = parse_salary_range(
                                                raw_text=raw_fb, min_text=min_fb, max_text=max_fb,
                                                default_currency=sal_cfg.get("default_currency", "SGD"),
                                                enable_period_inference=sal_cfg.get("enable_period_inference", True),
                                            )
                                            row["salary_raw"]      = fb_result.salary_raw
                                            row["salary_min"]      = fb_result.salary_min
                                            row["salary_max"]      = fb_result.salary_max
                                            row["salary_currency"] = fb_result.salary_currency
                                            row["salary_period"]   = fb_result.salary_period
                                            row["salary_status"]   = fb_result.salary_status
                                    except Exception:
                                        pass  # salary fallback is best-effort

                            except Exception as exc:
                                row["description_status"] = "ERROR"
                                row["notes"] = str(exc)
                                logger.warning("[careersfuture] Desc fetch error %s: %s", job_url, exc)
                            finally:
                                await desc_page.close()
                            _delay(delay_range)
                        else:
                            row["description_status"] = "SKIPPED"

                        append_row(out_path, row)
                        existing_urls.add(job_url)
                        added += 1
                        logger.info(
                            "[careersfuture] ✓ %s @ %s [%s]",
                            listing["role"], listing["company"], row["description_status"],
                        )

                except Exception as exc:
                    logger.error("[careersfuture] Page %d error: %s", page_num, exc)
                    try:
                        await search_page.close()
                    except Exception:
                        pass

    finally:
        await context.close()

    return added


# ---------------------------------------------------------------------------
# Delay helper
# ---------------------------------------------------------------------------

def _delay(delay_range: tuple[float, float]) -> None:
    wait = random.uniform(*delay_range)
    logger.debug("Waiting %.1fs...", wait)
    time.sleep(wait)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def _run(args: argparse.Namespace) -> None:
    with open(CONFIG_PATH) as f:
        config = yaml.safe_load(f)

    out_path = Path(args.output)
    init_excel(out_path)
    existing_urls = load_existing_urls(out_path)
    logger.info("Output: %s  |  Existing rows: %d", out_path, len(existing_urls))

    roles = load_roles(config)
    if not roles:
        logger.error("No roles found. Add roles to job_roles.xlsx or config.yaml.")
        return

    portals_cfg = config.get("portals", {})
    portal_filter = args.portal.lower() if args.portal else None

    total_added = 0

    async with async_playwright() as playwright:
        # CareersFuture
        if (portal_filter is None or portal_filter == "careersfuture") \
                and portals_cfg.get("careersfuture", {}).get("enabled", False):
            cfg = portals_cfg["careersfuture"]
            delay_range = (
                float(cfg.get("min_delay_seconds", 5)),
                float(cfg.get("max_delay_seconds", 15)),
            )
            n = await scrape_careersfuture(
                playwright=playwright,
                roles=roles,
                pages_per_role=args.pages,
                fetch_descriptions=not args.no_descriptions,
                out_path=out_path,
                existing_urls=existing_urls,
                delay_range=delay_range,
            )
            total_added += n
            logger.info("[careersfuture] Added %d new rows.", n)

        # Indeed — skip if selectors are not configured
        if (portal_filter is None or portal_filter == "indeed") \
                and portals_cfg.get("indeed", {}).get("enabled", False):
            logger.warning(
                "[indeed] Skipped — CSS selectors not yet configured in adapters/indeed.py."
            )

    logger.info("Trawl complete. Total new rows added: %d → %s", total_added, out_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Standalone job trawler — scrapes listings and logs to Excel."
    )
    parser.add_argument(
        "--pages", type=int, default=2,
        help="Number of result pages to scrape per role (default: 2)",
    )
    parser.add_argument(
        "--portal", type=str, default=None,
        help="Limit to a specific portal: careersfuture | indeed (default: all enabled)",
    )
    parser.add_argument(
        "--output", type=str, default=str(DEFAULT_OUT),
        help=f"Output Excel file path (default: {DEFAULT_OUT})",
    )
    parser.add_argument(
        "--no-descriptions", action="store_true",
        help="Skip fetching full job descriptions (faster, URL+title+company only)",
    )
    args = parser.parse_args()

    _check_browser_installed()

    logger.info(
        "Starting trawl | pages=%d | portal=%s | output=%s | descriptions=%s",
        args.pages,
        args.portal or "all",
        args.output,
        "no" if args.no_descriptions else "yes",
    )
    asyncio.run(_run(args))


if __name__ == "__main__":
    main()
