from pathlib import Path

from openpyxl import Workbook, load_workbook

from skills_filter_pipeline import enrich_trawl_results, ensure_skills_input_file


def _write_trawl_sample(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trawl"
    ws.append([
        "id", "scraped_at", "portal", "role", "company", "url", "page_num",
        "raw_description", "description_status", "notes",
    ])
    ws.append([
        "j1", "", "careersfuture", "Data Engineer", "Acme", "https://example.com/1", 1,
        "Must have SQL, Python, and Spark. Experience with Airflow preferred.",
        "OK", "",
    ])
    ws.append([
        "j2", "", "careersfuture", "Analyst", "Beta", "https://example.com/2", 1,
        "", "OK", "",
    ])
    wb.save(path)


def test_enrich_trawl_results_adds_columns_and_values(tmp_path: Path) -> None:
    f = tmp_path / "trawl.xlsx"
    skills_file = tmp_path / "input" / "skills_input.xlsx"
    _write_trawl_sample(f)
    ensure_skills_input_file(skills_file)

    updated, with_skills, saved = enrich_trawl_results(f, default_continue=0, skills_file=skills_file)

    assert updated == 2
    assert with_skills == 1
    assert saved == f

    wb = load_workbook(f)
    ws = wb.active
    header = {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}

    assert "skills" in header
    assert "continue" in header

    skills_row1 = str(ws.cell(row=2, column=header["skills"]).value or "")
    continue_row1 = ws.cell(row=2, column=header["continue"]).value

    assert "sql" in skills_row1
    assert "python" in skills_row1
    assert "spark" in skills_row1
    assert continue_row1 == 0


def test_ensure_skills_input_file_creates_excel(tmp_path: Path) -> None:
    skills_file = tmp_path / "input" / "skills_input.xlsx"
    ensure_skills_input_file(skills_file)

    assert skills_file.exists()
    wb = load_workbook(skills_file)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "skill"
    assert ws.cell(row=1, column=2).value == "regex"
    assert ws.max_row > 2
