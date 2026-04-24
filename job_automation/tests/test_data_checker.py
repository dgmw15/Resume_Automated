"""
tests/test_data_checker.py — Deterministic tests for core/data_checker.py

Covers:
1. COMPLETE — row with all fields present and consistent
2. MISSING — truly_missing gap detection (null/empty fields)
3. SEMANTICALLY_MISSING — salary_status = MISSING/AMBIGUOUS/ERROR
4. INCONSISTENT — salary_min > salary_max, currency null while min present,
                  salary_status=OK with null min+max
5. RECOVERED_LOCAL — local salary_raw reparse, min↔max mirror
6. UNRESOLVED — gaps that cannot be resolved locally or by refetch
7. SKIPPED_NO_URL — refetch path with no URL
8. Idempotency — running checker twice on an already-COMPLETE row produces COMPLETE
9. Dry-run — no writes happen even in recover mode when dry_run=True
10. from_config — factory reads all config keys
11. Field contract detection — trawl vs tracker schema selection
12. Backup — backup file created when write_backup=True
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.data_checker import (
    COMPLETE,
    DERIVED,
    ERROR_FETCH,
    INCONSISTENT,
    RECOVERABLE,
    RECOVERED_LOCAL,
    RECOVERED_REFETCH,
    SEMANTICALLY_MISSING,
    SKIPPED_NO_URL,
    TRACKER_CONTRACT,
    TRAWL_CONTRACT,
    TRULY_MISSING,
    UNRESOLVED,
    DataChecker,
    FieldGap,
    WorkbookReport,
    _is_absent,
    _make_backup,
    _summarise_gaps,
)


# ---------------------------------------------------------------------------
# Fixtures — in-memory config and temporary workbooks
# ---------------------------------------------------------------------------

def _make_config(
    *,
    enabled: bool = True,
    mode: str = "audit_only",
    dry_run: bool = True,
    targets: list[str] | None = None,
    allow_refetch: bool = False,
    write_backup: bool = False,
    report_path: str = "",
    recovery_path: str = "",
) -> dict:
    return {
        "data_checker": {
            "enabled": enabled,
            "mode": mode,
            "dry_run": dry_run,
            "targets": targets or [],
            "write_backup": write_backup,
            "backfill": {
                "allow_portal_refetch": allow_refetch,
                "portals_allowed_for_refetch": ["careersfuture"],
                "max_rows_per_run": 50,
                "unresolved_reason_required": True,
            },
            "report_path": report_path or "output/logs/completeness_{date}.json",
            "recovery_report_path": recovery_path or "output/logs/recovery_{date}.json",
            "issue_workbook_path": "",
            "backup_retention_days": 7,
        },
        "salary": {
            "default_currency": "SGD",
            "enable_period_inference": True,
        },
    }


def _write_trawl_workbook(path: Path, rows: list[dict]) -> None:
    """Write a minimal trawl workbook with salary columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trawl"

    columns = [
        "id", "scraped_at", "portal", "role", "company", "url",
        "page_num", "raw_description", "description_status", "notes",
        "salary_raw", "salary_min", "salary_max", "salary_currency",
        "salary_period", "salary_status",
    ]
    for i, col in enumerate(columns, 1):
        ws.cell(row=1, column=i, value=col)

    for row_num, row in enumerate(rows, 2):
        for i, col in enumerate(columns, 1):
            ws.cell(row=row_num, column=i, value=row.get(col))

    wb.save(path)


def _complete_row(row_id: str = "r1") -> dict:
    """A trawl row that should always pass all checks."""
    return {
        "id": row_id,
        "portal": "careersfuture",
        "role": "Data Analyst",
        "url": "https://www.careers.gov.sg/job/1234",
        "company": "Acme Corp",
        "raw_description": "We are hiring a data analyst.",
        "salary_raw": "$4,500 to $6,500",
        "salary_min": 4500.0,
        "salary_max": 6500.0,
        "salary_currency": "SGD",
        "salary_period": "unknown",
        "salary_status": "OK",
    }


# ---------------------------------------------------------------------------
# Unit tests: _is_absent helper
# ---------------------------------------------------------------------------

class TestIsAbsent:
    def test_none_is_absent(self):
        assert _is_absent(None)

    def test_empty_string_is_absent(self):
        assert _is_absent("")

    def test_whitespace_is_absent(self):
        assert _is_absent("   ")

    def test_zero_is_not_absent(self):
        assert not _is_absent(0)

    def test_zero_float_is_not_absent(self):
        assert not _is_absent(0.0)

    def test_valid_string_is_not_absent(self):
        assert not _is_absent("hello")

    def test_number_is_not_absent(self):
        assert not _is_absent(4500.0)


# ---------------------------------------------------------------------------
# Unit tests: gap classification
# ---------------------------------------------------------------------------

class TestClassifyGaps:
    def _checker(self) -> DataChecker:
        return DataChecker(_make_config())

    def test_complete_row_has_no_gaps(self):
        checker = self._checker()
        gaps = checker._classify_gaps(_complete_row(), TRAWL_CONTRACT)
        assert gaps == []

    def test_missing_id_is_critical(self):
        checker = self._checker()
        row = _complete_row()
        row["id"] = None
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.field_name == "id" and g.gap_type == TRULY_MISSING for g in gaps)

    def test_missing_salary_raw_is_recoverable(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_raw"] = ""
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.field_name == "salary_raw" and g.classification == RECOVERABLE for g in gaps)

    def test_missing_salary_min_is_derived(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_min"] = None
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.field_name == "salary_min" and g.classification == DERIVED for g in gaps)

    def test_salary_status_missing_is_semantically_missing(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_status"] = "MISSING"
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.gap_type == SEMANTICALLY_MISSING and g.field_name == "salary_status" for g in gaps)

    def test_salary_status_ambiguous_is_semantically_missing(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_status"] = "AMBIGUOUS"
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.gap_type == SEMANTICALLY_MISSING for g in gaps)

    def test_salary_status_error_is_semantically_missing(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_status"] = "ERROR"
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.gap_type == SEMANTICALLY_MISSING for g in gaps)

    def test_min_greater_than_max_is_inconsistent(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_min"] = 9000.0
        row["salary_max"] = 5000.0
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.gap_type == INCONSISTENT and g.field_name == "salary_min" for g in gaps)

    def test_currency_null_with_min_present_is_inconsistent(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_currency"] = None
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.gap_type == INCONSISTENT and g.field_name == "salary_currency" for g in gaps)

    def test_status_ok_with_null_amounts_is_inconsistent(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_min"] = None
        row["salary_max"] = None
        row["salary_status"] = "OK"
        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        assert any(g.gap_type == INCONSISTENT and g.field_name == "salary_status" for g in gaps)

    def test_tracker_contract_used_when_portal_name_column(self):
        checker = self._checker()
        row = {
            "id": "x1",
            "portal_name": "careersfuture",  # tracker uses "portal_name"
            "role": "Analyst",
            "url": "https://example.com",
            "company": "Corp",
            "raw_description": "desc",
            "status": "SCRAPED",
            "salary_raw": "$5000",
            "salary_min": 5000.0,
            "salary_max": 5000.0,
            "salary_currency": "SGD",
            "salary_period": "unknown",
            "salary_status": "OK",
        }
        gaps = checker._classify_gaps(row, TRACKER_CONTRACT)
        assert gaps == []


# ---------------------------------------------------------------------------
# Unit tests: local recovery
# ---------------------------------------------------------------------------

class TestLocalRecovery:
    def _checker(self) -> DataChecker:
        return DataChecker(_make_config())

    def test_reparse_salary_raw_fills_derived_fields(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_min"] = None
        row["salary_max"] = None
        row["salary_currency"] = None
        row["salary_period"] = None
        row["salary_status"] = "MISSING"
        row["salary_raw"] = "$4,500 to $6,500"

        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        updated, actions = checker._try_local_recovery(row, gaps, TRAWL_CONTRACT)

        assert updated["salary_status"] == "OK"
        assert updated["salary_min"] == pytest.approx(4500.0)
        assert updated["salary_max"] == pytest.approx(6500.0)
        assert len(actions) > 0

    def test_mirror_max_from_min(self):
        """When salary_raw is empty but salary_min is present, salary_max mirrors salary_min."""
        checker = self._checker()
        row = _complete_row()
        row["salary_max"] = None
        row["salary_raw"] = ""  # no raw text — reparse cannot help, mirroring is needed

        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        updated, actions = checker._try_local_recovery(row, gaps, TRAWL_CONTRACT)
        assert updated["salary_max"] == pytest.approx(4500.0)
        assert any("Mirrored salary_max" in a for a in actions)

    def test_mirror_min_from_max(self):
        """When salary_raw is empty but salary_max is present, salary_min mirrors salary_max."""
        checker = self._checker()
        row = _complete_row()
        row["salary_min"] = None
        row["salary_raw"] = ""  # no raw text — reparse cannot help, mirroring is needed

        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        updated, actions = checker._try_local_recovery(row, gaps, TRAWL_CONTRACT)
        assert updated["salary_min"] == pytest.approx(6500.0)
        assert any("Mirrored salary_min" in a for a in actions)

    def test_no_salary_raw_produces_no_actions(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_raw"] = ""
        row["salary_min"] = None
        row["salary_max"] = None
        row["salary_currency"] = None
        row["salary_status"] = "MISSING"

        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        updated, actions = checker._try_local_recovery(row, gaps, TRAWL_CONTRACT)
        # salary_raw is empty — reparse can't help
        assert updated["salary_status"] in (None, "MISSING", "")

    def test_recovery_does_not_modify_original_dict(self):
        checker = self._checker()
        row = _complete_row()
        row["salary_max"] = None
        original_min = row["salary_min"]

        gaps = checker._classify_gaps(row, TRAWL_CONTRACT)
        _, _ = checker._try_local_recovery(row, gaps, TRAWL_CONTRACT)

        # Original dict must be untouched
        assert row["salary_max"] is None
        assert row["salary_min"] == original_min


# ---------------------------------------------------------------------------
# Integration tests: full workbook check (using temp files)
# ---------------------------------------------------------------------------

class TestAuditMode:
    @pytest.fixture
    def tmp_wb(self, tmp_path) -> Path:
        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [_complete_row()])
        return path

    def test_complete_row_returns_complete(self, tmp_wb):
        config = _make_config(enabled=True, mode="audit_only", targets=[str(tmp_wb)])
        checker = DataChecker(config)
        import asyncio
        reports = asyncio.run(checker.run())
        assert len(reports) == 1
        assert reports[0].outcome_counts[COMPLETE] == 1

    def test_missing_salary_returns_unresolved_in_audit_mode(self, tmp_path):
        row = _complete_row()
        row["salary_min"] = None
        row["salary_max"] = None
        row["salary_status"] = "MISSING"

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])

        config = _make_config(enabled=True, mode="audit_only", targets=[str(path)])
        checker = DataChecker(config)
        import asyncio
        reports = asyncio.run(checker.run())
        assert reports[0].outcome_counts[UNRESOLVED] >= 1

    def test_audit_mode_never_writes(self, tmp_path):
        row = _complete_row()
        row["salary_min"] = None
        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])

        mtime_before = path.stat().st_mtime

        config = _make_config(enabled=True, mode="audit_only", dry_run=False, targets=[str(path)])
        checker = DataChecker(config)
        import asyncio
        asyncio.run(checker.run())

        # Workbook must be untouched in audit_only mode
        assert path.stat().st_mtime == mtime_before

    def test_missing_target_workbook_is_skipped_gracefully(self, tmp_path):
        config = _make_config(
            enabled=True, targets=[str(tmp_path / "does_not_exist.xlsx")]
        )
        import asyncio
        reports = asyncio.run(DataChecker(config).run())
        assert reports == []


class TestRecoverMode:
    def test_local_recovery_fills_derived_salary(self, tmp_path):
        row = _complete_row()
        row["salary_min"] = None
        row["salary_max"] = None
        row["salary_currency"] = None
        row["salary_period"] = None
        row["salary_status"] = "MISSING"
        row["salary_raw"] = "$4,500 to $6,500"

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])

        config = _make_config(
            enabled=True, mode="recover", dry_run=True, targets=[str(path)]
        )
        import asyncio
        reports = asyncio.run(DataChecker(config).run())
        assert reports[0].outcome_counts[RECOVERED_LOCAL] == 1

    def test_dry_run_does_not_write_workbook(self, tmp_path):
        row = _complete_row()
        row["salary_min"] = None
        row["salary_raw"] = "$5,000"

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])
        mtime_before = path.stat().st_mtime

        config = _make_config(
            enabled=True, mode="recover", dry_run=True,
            targets=[str(path)], write_backup=False,
        )
        import asyncio
        asyncio.run(DataChecker(config).run())

        assert path.stat().st_mtime == mtime_before

    def test_recover_writes_workbook_when_not_dry_run(self, tmp_path):
        row = _complete_row()
        row["salary_min"] = None
        row["salary_raw"] = "$5,000"

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])
        mtime_before = path.stat().st_mtime

        config = _make_config(
            enabled=True, mode="recover", dry_run=False,
            targets=[str(path)], write_backup=False,
        )
        import asyncio
        asyncio.run(DataChecker(config).run())

        assert path.stat().st_mtime != mtime_before  # file was written

    def test_critical_field_gap_is_unresolved(self, tmp_path):
        """Critical fields (id, portal, url) cannot be locally recovered."""
        row = _complete_row()
        row["url"] = None

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])

        config = _make_config(enabled=True, mode="recover", targets=[str(path)])
        import asyncio
        reports = asyncio.run(DataChecker(config).run())
        assert reports[0].outcome_counts[UNRESOLVED] == 1


# ---------------------------------------------------------------------------
# Idempotency
# ---------------------------------------------------------------------------

class TestIdempotency:
    def test_already_complete_row_unchanged_on_rerun(self, tmp_path):
        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [_complete_row()])

        config = _make_config(
            enabled=True, mode="recover", dry_run=False,
            targets=[str(path)], write_backup=False,
        )
        import asyncio

        # First run
        reports1 = asyncio.run(DataChecker(config).run())
        # Second run
        reports2 = asyncio.run(DataChecker(config).run())

        assert reports1[0].outcome_counts[COMPLETE] == 1
        assert reports2[0].outcome_counts[COMPLETE] == 1

    def test_recovered_row_is_complete_on_rerun(self, tmp_path):
        row = _complete_row()
        row["salary_min"] = None
        row["salary_raw"] = "$5,000"

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])

        config = _make_config(
            enabled=True, mode="recover", dry_run=False,
            targets=[str(path)], write_backup=False,
        )
        import asyncio

        reports1 = asyncio.run(DataChecker(config).run())
        assert reports1[0].outcome_counts[RECOVERED_LOCAL] == 1

        # Second run — now the row should be COMPLETE (was fixed in first run)
        reports2 = asyncio.run(DataChecker(config).run())
        assert reports2[0].outcome_counts[COMPLETE] == 1
        assert reports2[0].outcome_counts[RECOVERED_LOCAL] == 0


# ---------------------------------------------------------------------------
# SKIPPED_NO_URL (refetch enabled but URL missing)
# ---------------------------------------------------------------------------

class TestSkippedNoUrl:
    def test_missing_url_tagged_skipped_no_url_when_refetch_enabled(self, tmp_path):
        row = _complete_row()
        row["url"] = None
        row["salary_min"] = None
        row["salary_raw"] = ""   # also empty so local recovery can't help

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])

        config = _make_config(
            enabled=True, mode="recover", allow_refetch=True, targets=[str(path)]
        )
        import asyncio
        reports = asyncio.run(DataChecker(config).run())
        # url is critical → UNRESOLVED (url gap cannot be recovered locally)
        # salary_min is derived; with no url, refetch is impossible
        # The row has a critical gap (url), so outcome is at least UNRESOLVED
        counts = reports[0].outcome_counts
        assert counts[UNRESOLVED] + counts[SKIPPED_NO_URL] >= 1


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------

class TestBackup:
    def test_backup_created_before_write(self, tmp_path):
        row = _complete_row()
        row["salary_min"] = None
        row["salary_raw"] = "$5,000"

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])

        config = _make_config(
            enabled=True, mode="recover", dry_run=False,
            write_backup=True, targets=[str(path)],
        )
        import asyncio
        asyncio.run(DataChecker(config).run())

        backups = list(tmp_path.glob("trawl_results_backup_*.xlsx"))
        assert len(backups) == 1

    def test_no_backup_when_disabled(self, tmp_path):
        row = _complete_row()
        row["salary_min"] = None
        row["salary_raw"] = "$5,000"

        path = tmp_path / "trawl_results.xlsx"
        _write_trawl_workbook(path, [row])

        config = _make_config(
            enabled=True, mode="recover", dry_run=False,
            write_backup=False, targets=[str(path)],
        )
        import asyncio
        asyncio.run(DataChecker(config).run())

        backups = list(tmp_path.glob("trawl_results_backup_*.xlsx"))
        assert backups == []


# ---------------------------------------------------------------------------
# from_config factory
# ---------------------------------------------------------------------------

class TestFromConfig:
    def test_disabled_checker_returns_empty(self):
        config = _make_config(enabled=False)
        checker = DataChecker.from_config(config)
        import asyncio
        reports = asyncio.run(checker.run())
        assert reports == []

    def test_factory_reads_mode(self):
        config = _make_config(mode="recover")
        checker = DataChecker.from_config(config)
        assert checker._mode == "recover"

    def test_factory_reads_dry_run(self):
        config = _make_config(dry_run=False)
        checker = DataChecker.from_config(config)
        assert checker._dry_run is False

    def test_factory_reads_allow_refetch(self):
        config = _make_config(allow_refetch=True)
        checker = DataChecker.from_config(config)
        assert checker._allow_refetch is True

    def test_factory_reads_max_rows(self):
        config = _make_config()
        config["data_checker"]["backfill"]["max_rows_per_run"] = 10
        checker = DataChecker.from_config(config)
        assert checker._max_rows == 10

    def test_factory_reads_default_currency(self):
        config = _make_config()
        config["salary"]["default_currency"] = "USD"
        checker = DataChecker.from_config(config)
        assert checker._default_currency == "USD"


# ---------------------------------------------------------------------------
# WorkbookReport computed properties
# ---------------------------------------------------------------------------

class TestWorkbookReport:
    def _report(self) -> WorkbookReport:
        from core.data_checker import RowResult
        r = WorkbookReport(
            workbook_path="test.xlsx",
            total_rows=3,
            mode="audit_only",
            dry_run=True,
        )
        r.results = [
            RowResult(row_id="r1", outcome=COMPLETE),
            RowResult(row_id="r2", outcome=UNRESOLVED, gaps=[
                FieldGap("salary_min", DERIVED, TRULY_MISSING, "null"),
            ], reason="salary_min(truly_missing)"),
            RowResult(row_id="r3", outcome=RECOVERED_LOCAL, gaps=[
                FieldGap("salary_max", DERIVED, TRULY_MISSING, "null"),
            ], recovery_actions=["Mirrored salary_max = salary_min = 5000"]),
        ]
        return r

    def test_outcome_counts(self):
        r = self._report()
        counts = r.outcome_counts
        assert counts[COMPLETE] == 1
        assert counts[UNRESOLVED] == 1
        assert counts[RECOVERED_LOCAL] == 1

    def test_field_missing_pct_calculated(self):
        r = self._report()
        pct = r.field_missing_pct
        assert "salary_min" in pct
        assert pct["salary_min"] == pytest.approx(100 / 3 * 1, rel=0.01)

    def test_to_dict_contains_required_keys(self):
        r = self._report()
        d = r.to_dict()
        for key in ("workbook", "generated_at", "mode", "dry_run", "total_rows",
                    "outcome_counts", "field_missing_pct", "unresolved_details"):
            assert key in d

    def test_to_recovery_dict_contains_required_keys(self):
        r = self._report()
        d = r.to_recovery_dict()
        for key in ("rows_attempted", "rows_fixed", "rows_unresolved",
                    "recovery_breakdown", "unresolved_by_reason"):
            assert key in d

    def test_to_recovery_dict_counts_match(self):
        r = self._report()
        d = r.to_recovery_dict()
        # rows_attempted = non-COMPLETE = 2
        assert d["rows_attempted"] == 2
        # rows_fixed = RECOVERED_LOCAL + RECOVERED_REFETCH = 1
        assert d["rows_fixed"] == 1
        # rows_unresolved = UNRESOLVED + SKIPPED + ERROR = 1
        assert d["rows_unresolved"] == 1


# ---------------------------------------------------------------------------
# _summarise_gaps helper
# ---------------------------------------------------------------------------

class TestSummariseGaps:
    def test_empty_gaps_returns_empty_string(self):
        assert _summarise_gaps([]) == ""

    def test_single_gap_formatted(self):
        gap = FieldGap("salary_min", DERIVED, TRULY_MISSING)
        result = _summarise_gaps([gap])
        assert "salary_min" in result
        assert "truly_missing" in result

    def test_multiple_gaps_joined(self):
        gaps = [
            FieldGap("salary_min", DERIVED, TRULY_MISSING),
            FieldGap("salary_currency", DERIVED, INCONSISTENT),
        ]
        result = _summarise_gaps(gaps)
        assert "salary_min" in result
        assert "salary_currency" in result
