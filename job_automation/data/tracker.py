from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from data.models import JobListing, JobStatus

DB_PATH = Path("Database.xlsx")

COLUMNS = [
    "id", "portal_name", "role", "company", "url",
    "raw_description", "tailored_resume", "status", "page_num", "timestamp",
    "validation_score", "validation_reason", "pipeline_track",
    "ai_provider_used", "cost_usd", "cost_reserved_usd", "cost_actual_usd",
    "reservation_id", "reservation_expires_at", "idempotency_key",
    "docx_path", "docx_validation_error", "processed_at",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class ExcelTracker:
    def __init__(self, path: Path = DB_PATH) -> None:
        self.path = path
        self._init_workbook()

    # ------------------------------------------------------------------
    # Initialisation / migration
    # ------------------------------------------------------------------

    def _init_workbook(self) -> None:
        if not self.path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Jobs"
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            wb.save(self.path)
        else:
            # Migrate: add any missing columns to an existing workbook
            self._migrate_columns()

    def _migrate_columns(self) -> None:
        """Add new columns to existing workbook without touching existing data."""
        wb = load_workbook(self.path)
        ws = wb["Jobs"]
        existing = {ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)}
        for col_name in COLUMNS:
            if col_name not in existing:
                next_col = ws.max_column + 1
                cell = ws.cell(row=1, column=next_col, value=col_name)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
        wb.save(self.path)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _load(self):
        return load_workbook(self.path)

    def _col_map(self, ws) -> dict[str, int]:
        """Return {column_name: column_index} from row 1."""
        return {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

    def _set_cell(self, ws, row_idx: int, col: dict[str, int], name: str, value) -> None:
        """Write value to cell only if the column exists in this workbook."""
        if name in col:
            ws.cell(row=row_idx, column=col[name]).value = value

    # ------------------------------------------------------------------
    # Public API — existing
    # ------------------------------------------------------------------

    def append(self, job: JobListing) -> None:
        """Append a new JobListing row. Skips if URL already exists."""
        wb = self._load()
        ws = wb["Jobs"]
        col = self._col_map(ws)

        # Deduplicate by URL
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[col["url"] - 1] == job.url:
                return

        next_row = ws.max_row + 1
        data = job.model_dump()
        for name, idx in col.items():
            value = data.get(name)
            if isinstance(value, datetime):
                value = value.isoformat()
            elif hasattr(value, "value"):  # Enum
                value = value.value
            ws.cell(row=next_row, column=idx, value=value)

        wb.save(self.path)

    def update(self, job_id: str, status: Optional[JobStatus] = None,
               tailored_resume: Optional[str] = None, **kwargs) -> bool:
        """Update status, tailored_resume, and/or arbitrary extra columns by job_id."""
        wb = self._load()
        ws = wb["Jobs"]
        col = self._col_map(ws)

        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=col["id"]).value == job_id:
                if status is not None:
                    self._set_cell(ws, row_idx, col, "status", status.value)
                if tailored_resume is not None:
                    self._set_cell(ws, row_idx, col, "tailored_resume", tailored_resume)
                for k, v in kwargs.items():
                    if isinstance(v, datetime):
                        v = v.isoformat()
                    elif hasattr(v, "value"):
                        v = v.value
                    self._set_cell(ws, row_idx, col, k, v)
                wb.save(self.path)
                return True
        return False

    def update_by_url(self, url: str, status: Optional[JobStatus] = None,
                      tailored_resume: Optional[str] = None) -> bool:
        """Update status and/or tailored_resume for a row matching URL."""
        wb = self._load()
        ws = wb["Jobs"]
        col = self._col_map(ws)

        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=col["url"]).value == url:
                if status is not None:
                    self._set_cell(ws, row_idx, col, "status", status.value)
                if tailored_resume is not None:
                    self._set_cell(ws, row_idx, col, "tailored_resume", tailored_resume)
                wb.save(self.path)
                return True
        return False

    def get_last_page(self, portal_name: str) -> int:
        """Return the highest page_num scraped for a given portal (0 if none)."""
        wb = self._load()
        ws = wb["Jobs"]
        col = self._col_map(ws)

        max_page = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            portal_val = row[col["portal_name"] - 1]
            page_val = row[col["page_num"] - 1]
            if portal_val == portal_name and isinstance(page_val, (int, float)):
                max_page = max(max_page, int(page_val))
        return max_page

    def get_by_status(self, status: JobStatus) -> list[dict]:
        """Return all rows matching a given status as list of dicts."""
        return self.list_rows_by_status(status)

    # ------------------------------------------------------------------
    # Public API — new helpers
    # ------------------------------------------------------------------

    def list_rows_by_status(self, status: JobStatus) -> list[dict]:
        """Return all rows matching a given status as list of dicts."""
        wb = self._load()
        ws = wb["Jobs"]
        col = self._col_map(ws)
        col_names = list(col.keys())

        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {col_names[i]: row[i] for i in range(len(col_names))}
            if row_dict.get("status") == status.value:
                results.append(row_dict)
        return results

    def mark_validation_result(
        self,
        job_id: str,
        status: JobStatus,
        score: int,
        reason: str,
    ) -> bool:
        return self.update(
            job_id,
            status=status,
            validation_score=score,
            validation_reason=reason,
        )

    def mark_batch_queued(self, job_id: str) -> bool:
        return self.update(job_id, status=JobStatus.BATCH_QUEUED)

    def mark_ai_result(
        self,
        job_id: str,
        tailored_text: str,
        provider: str,
        cost_usd: float,
        pipeline_track: str,
    ) -> bool:
        return self.update(
            job_id,
            status=JobStatus.TAILORED_TEXT_READY,
            tailored_resume=tailored_text,
            ai_provider_used=provider,
            cost_usd=cost_usd,
            pipeline_track=pipeline_track,
            processed_at=datetime.utcnow().isoformat(),
        )

    def mark_docx_ready(self, job_id: str, docx_path: str) -> bool:
        return self.update(job_id, status=JobStatus.DOCX_READY, docx_path=docx_path)

    def mark_docx_failed(self, job_id: str, error: str) -> bool:
        return self.update(
            job_id,
            status=JobStatus.DOCX_GENERATION_FAILED,
            docx_validation_error=error,
        )

    def mark_reservation(
        self,
        job_id: str,
        reservation_id: str,
        reserved_usd: float,
        expires_at: datetime,
        idempotency_key: str,
    ) -> bool:
        return self.update(
            job_id,
            reservation_id=reservation_id,
            cost_reserved_usd=reserved_usd,
            reservation_expires_at=expires_at.isoformat(),
            idempotency_key=idempotency_key,
        )

    def mark_reservation_committed(
        self,
        job_id: str,
        actual_usd: float,
    ) -> bool:
        return self.update(
            job_id,
            cost_actual_usd=actual_usd,
            cost_usd=actual_usd,
            reservation_id=None,
            reservation_expires_at=None,
        )

    def get_active_reservations(self) -> list[dict]:
        """Return all rows that currently hold a non-null reservation_id."""
        wb = self._load()
        ws = wb["Jobs"]
        col = self._col_map(ws)
        col_names = list(col.keys())

        results = []
        if "reservation_id" not in col:
            return results
        rid_idx = col["reservation_id"] - 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[rid_idx] not in (None, ""):
                results.append({col_names[i]: row[i] for i in range(len(col_names))})
        return results
