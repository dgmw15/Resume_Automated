"""Populate skills + continue columns in trawl_results.xlsx.

Usage:
    python skills_filter_pipeline.py
    python skills_filter_pipeline.py --input trawl_results.xlsx --default-continue 0
    python skills_filter_pipeline.py --skills-file input/skills_input.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ai.skills_signal_extractor import DEFAULT_SKILL_PATTERNS, extract_technical_skills

DEFAULT_INPUT = Path("trawl_results.xlsx")
INPUT_DIR = Path("input")
DEFAULT_SKILLS_FILE = INPUT_DIR / "skills_input.xlsx"
REQUIRED_COLUMNS = ["skills", "continue"]


def ensure_skills_input_file(path: Path) -> None:
    if path.exists():
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(["skill", "regex", "enabled", "notes"])
    for skill, regex in DEFAULT_SKILL_PATTERNS:
        ws.append([skill, regex, 1, ""])
    wb.save(path)


def load_skill_patterns_from_excel(path: Path) -> list[tuple[str, str]]:
    ensure_skills_input_file(path)
    wb = load_workbook(path)
    ws = wb.active

    header = {str(ws.cell(row=1, column=i).value).strip().lower(): i for i in range(1, ws.max_column + 1)}
    skill_col = header.get("skill")
    regex_col = header.get("regex")
    enabled_col = header.get("enabled")

    if not skill_col:
        return DEFAULT_SKILL_PATTERNS

    patterns: list[tuple[str, str]] = []
    for r in range(2, ws.max_row + 1):
        skill = str(ws.cell(row=r, column=skill_col).value or "").strip()
        if not skill:
            continue

        enabled = 1
        if enabled_col:
            raw_enabled = str(ws.cell(row=r, column=enabled_col).value or "1").strip().lower()
            if raw_enabled in {"0", "false", "no", "n"}:
                enabled = 0

        if not enabled:
            continue

        regex = ""
        if regex_col:
            regex = str(ws.cell(row=r, column=regex_col).value or "").strip()
        if not regex:
            regex = rf"\b{skill.lower().replace(' ', r'\s*')}\b"

        patterns.append((skill.lower(), regex))

    return patterns or DEFAULT_SKILL_PATTERNS


def _ensure_columns(ws) -> dict[str, int]:
    header = {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}
    for col in REQUIRED_COLUMNS:
        if col not in header:
            idx = ws.max_column + 1
            ws.cell(row=1, column=idx, value=col)
            header[col] = idx
    return header


def enrich_trawl_results(
    input_path: Path,
    default_continue: int = 0,
    output_path: Path | None = None,
    skills_file: Path = DEFAULT_SKILLS_FILE,
) -> tuple[int, int, Path]:
    skill_patterns = load_skill_patterns_from_excel(skills_file)
    wb = load_workbook(input_path)
    ws = wb.active
    header = _ensure_columns(ws)

    rows_updated = 0
    rows_with_skills = 0

    for r in range(2, ws.max_row + 1):
        desc_col = header.get("raw_description")
        if not desc_col:
            continue

        description = str(ws.cell(row=r, column=desc_col).value or "").strip()
        skills = extract_technical_skills(description, skill_patterns=skill_patterns)
        ws.cell(row=r, column=header["skills"], value=", ".join(skills))
        rows_updated += 1
        if skills:
            rows_with_skills += 1

        cont_val = ws.cell(row=r, column=header["continue"]).value
        if cont_val is None or str(cont_val).strip() == "":
            ws.cell(row=r, column=header["continue"], value=int(default_continue))

    save_path = output_path or input_path
    wb.save(save_path)
    return rows_updated, rows_with_skills, save_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Enrich trawl results with skills and continue columns.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input trawl results xlsx")
    parser.add_argument("--output", type=Path, default=None, help="Optional output xlsx path")
    parser.add_argument("--skills-file", type=Path, default=DEFAULT_SKILLS_FILE, help="Skills config xlsx")
    parser.add_argument("--default-continue", type=int, default=0, choices=[0, 1], help="Default continue value")
    args = parser.parse_args()

    updated, with_skills, saved = enrich_trawl_results(
        args.input,
        default_continue=args.default_continue,
        output_path=args.output,
        skills_file=args.skills_file,
    )
    print(
        f"Skills filter complete: rows_updated={updated}, rows_with_skills={with_skills}, "
        f"file={saved}, skills_file={args.skills_file}"
    )


if __name__ == "__main__":
    main()
