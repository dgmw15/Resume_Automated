"""
run_pipeline_sample.py — Run the real AI pipeline against N already-scraped
listings from trawl_results.xlsx, without needing a live browser/login.

Mirrors core/orchestrator.py's three phases (employment filter -> JD
validation -> AI batch) exactly, but scoped to a fixed sample pulled from
trawl output instead of a live scrape. Useful for sampling how the pipeline
behaves across many jobs at once — e.g. checking whether a WEAK ATS verdict
on one job is a one-off or a pattern — without repeated SingPass logins.

Writes real rows to Database.xlsx and real DOCX files, exactly like the
live pipeline would; this is not a dry run.

Usage:
    .venv\\Scripts\\activate
    python run_pipeline_sample.py --count 5
    python run_pipeline_sample.py --count 10 --input trawl_results.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import logging
from pathlib import Path

import openpyxl
import yaml

from ai.critic import AtsCritic
from ai.employment_filter import EmploymentFilter
from ai.jd_validator import build_validator_from_config
from ai.provider_router import ProviderRouter
from ai.tailor import ResumeTailor
from core.batch_processor import BatchProcessor, BASE_RESUME_PATH
from data.models import JobListing, JobStatus
from data.tracker import ExcelTracker
from output.docx_renderer import build_renderer_from_config

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("pipeline_sample")

CONFIG_PATH = Path("config.yaml")


def _load_candidates(input_path: Path, count: int, existing_urls: set[str]) -> list[dict]:
    wb = openpyxl.load_workbook(input_path, read_only=True)
    ws = wb["Trawl"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    candidates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("description_status") != "OK":
            continue
        if d.get("url") in existing_urls:
            continue
        candidates.append(d)
        if len(candidates) >= count:
            break
    return candidates


def _existing_urls(tracker: ExcelTracker) -> set[str]:
    if not tracker.path.exists():
        return set()
    wb = openpyxl.load_workbook(tracker.path, read_only=True)
    ws = wb["Jobs"]
    col = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}
    url_idx = col["url"] - 1
    return {r[url_idx] for r in ws.iter_rows(min_row=2, values_only=True) if r[url_idx]}


async def main(count: int, input_path: Path) -> None:
    if not BASE_RESUME_PATH.exists():
        logger.error("base_resume.txt not found — nothing to tailor with.")
        return

    config = yaml.safe_load(CONFIG_PATH.read_text())
    tracker = ExcelTracker()

    candidates = _load_candidates(input_path, count, _existing_urls(tracker))
    if not candidates:
        logger.error("No unprocessed OK-status rows found in %s.", input_path)
        return
    logger.info("Sampling %d listing(s) from %s.", len(candidates), input_path)

    job_ids: list[str] = []
    for c in candidates:
        job = JobListing(
            portal_name=c.get("portal") or "careersfuture",
            role=c.get("role") or "",
            company=c.get("company") or "",
            url=c.get("url") or "",
            raw_description=c.get("raw_description") or "",
            status=JobStatus.SCRAPED,
            page_num=c.get("page_num") or 1,
            salary_raw=c.get("salary_raw"),
            salary_min=c.get("salary_min"),
            salary_max=c.get("salary_max"),
            salary_currency=c.get("salary_currency"),
            salary_period=c.get("salary_period"),
            salary_status=c.get("salary_status"),
        )
        tracker.append(job)
        job_ids.append(job.id)

    # --- Phase 1: employment filter (mirrors orchestrator._phase_employment_filter) ---
    ef_cfg = config.get("employment_filter", {})
    if ef_cfg:
        emp_filter = EmploymentFilter.from_config(config)
        for row in tracker.get_by_status(JobStatus.SCRAPED):
            if row["id"] not in job_ids:
                continue
            result = emp_filter.classify(title=row.get("role", ""), description=row.get("raw_description", ""))
            tracker.mark_employment_filter(
                job_id=row["id"], status=result.status, reason=result.reason,
                emp_type_raw=result.employment_type_raw,
                emp_type_normalized=result.employment_type_normalized,
            )

    # --- Phase 2: JD validation (mirrors orchestrator._phase_validate) ---
    default_role = config.get("ai", {}).get("default_role", "analyst")
    validator = build_validator_from_config(config, role=default_role)
    for row in tracker.get_by_status(JobStatus.SCRAPED):
        if row["id"] not in job_ids or row.get("employment_filter_status") == "FILTERED":
            continue
        result = validator.validate(row.get("raw_description") or "")
        tracker.mark_validation_result(
            row["id"],
            status=JobStatus.VALIDATION_PASSED if result.is_pass else JobStatus.VALIDATION_FAILED_NON_TECH,
            score=result.score, reason=result.reason,
        )

    # --- Phase 3: AI tailoring for everything that passed ---
    router = ProviderRouter(config)
    tailor = ResumeTailor(router)
    critic = AtsCritic(router) if config.get("resume_tailoring", {}).get("enable_ats_critic", True) else None
    docx_renderer = build_renderer_from_config(config)
    batch_processor = BatchProcessor(tracker, tailor, config.get("batch", {}), docx_renderer=docx_renderer, critic=critic)
    base_resume = BASE_RESUME_PATH.read_text(encoding="utf-8")

    for row in tracker.get_by_status(JobStatus.VALIDATION_PASSED):
        if row["id"] not in job_ids:
            continue
        tracker.mark_batch_queued(row["id"])
        queued_row = next(r for r in tracker.get_by_status(JobStatus.BATCH_QUEUED) if r["id"] == row["id"])
        await batch_processor._process_row(queued_row, base_resume)

    # --- Summary ---
    print()
    print("=" * 100)
    print(f"{'ROLE':<35} {'COMPANY':<25} {'STATUS':<22} {'COVERAGE':<9} {'VERDICT':<8}")
    print("=" * 100)
    for job_id in job_ids:
        # Re-read final state for each job regardless of which status bucket it ended in
        row = None
        for status in JobStatus:
            matches = [r for r in tracker.get_by_status(status) if r["id"] == job_id]
            if matches:
                row = matches[0]
                break
        if row is None:
            continue
        role = (row.get("role") or "")[:34]
        company = (row.get("company") or "")[:24]
        status = str(row.get("status") or "")
        if row.get("employment_filter_status") == "FILTERED":
            status = f"FILTERED({row.get('employment_filter_reason', '')[:12]})"
        elif row.get("status") == "VALIDATION_FAILED_NON_TECH":
            status = f"VAL_FAILED(score={row.get('validation_score')})"
        coverage = row.get("keyword_coverage_score")
        coverage_s = f"{coverage}%" if coverage is not None else "-"
        verdict = row.get("ats_verdict") or "-"
        print(f"{role:<35} {company:<25} {status:<22} {coverage_s:<9} {verdict:<8}")
    print("=" * 100)
    print(f"Full detail in Database.xlsx and any generated DOCX files under output/docs/.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--count", type=int, default=5, help="Number of listings to sample (default: 5)")
    parser.add_argument("--input", type=str, default="trawl_results.xlsx", help="Source trawl workbook")
    args = parser.parse_args()
    asyncio.run(main(args.count, Path(args.input)))
