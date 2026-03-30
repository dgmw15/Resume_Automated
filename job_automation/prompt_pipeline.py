"""Build extraction-to-prompt payloads from trawl_results.xlsx.

Usage:
    python prompt_pipeline.py
    python prompt_pipeline.py --input trawl_results.xlsx --output output/prompts/prompts.jsonl
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from ai.pipeline import get_prompts, select_track

DEFAULT_INPUT = Path("trawl_results.xlsx")
DEFAULT_OUTPUT_DIR = Path("output/prompts")
DEFAULT_RESUME_PATH = Path("base_resume.txt")


def _read_rows(input_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(input_path)
    ws = wb.active
    header = {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}

    rows: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        rows.append({
            "id": str(ws.cell(row=r, column=header.get("id", 1)).value or ""),
            "role": str(ws.cell(row=r, column=header.get("role", 1)).value or ""),
            "raw_description": str(ws.cell(row=r, column=header.get("raw_description", 1)).value or ""),
            "description_status": str(ws.cell(row=r, column=header.get("description_status", 1)).value or ""),
            "skills": str(ws.cell(row=r, column=header.get("skills", 1)).value or ""),
            "continue": str(ws.cell(row=r, column=header.get("continue", 1)).value or ""),
        })
    return rows


def _build_user_prompt(user_template: str, description: str, base_resume: str) -> str:
    return user_template.format(job_description=description, base_resume=base_resume)


def build_prompt_records(input_path: Path, resume_path: Path, output_path: Path) -> tuple[int, int]:
    rows = _read_rows(input_path)

    if resume_path.exists():
        base_resume = resume_path.read_text(encoding="utf-8")
    else:
        base_resume = ""
        print(f"Warning: base resume not found at {resume_path}. Proceeding with empty resume block.")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    written = 0
    skipped = 0
    with output_path.open("w", encoding="utf-8") as f:
        for row in rows:
            desc = (row.get("raw_description") or "").strip()
            if not desc:
                skipped += 1
                continue

            status = (row.get("description_status") or "").upper()
            if status and status not in {"OK", "PENDING", ""}:
                skipped += 1
                continue

            continue_value = (row.get("continue") or "").strip().lower()
            if continue_value not in {"1", "true", "yes", "y"}:
                skipped += 1
                continue

            track = select_track(row.get("role", ""))
            system_prompt, user_template = get_prompts(track)

            user_prompt = _build_user_prompt(
                user_template=user_template,
                description=desc,
                base_resume=base_resume,
            )

            record = {
                "id": row.get("id", ""),
                "role": row.get("role", ""),
                "track": track,
                "system_prompt": system_prompt.strip(),
                "user_prompt": user_prompt.strip(),
            }
            f.write(json.dumps(record, ensure_ascii=True) + "\n")
            written += 1

    return written, skipped


def _default_output_path() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return DEFAULT_OUTPUT_DIR / f"prompts_{timestamp}.jsonl"


def main() -> None:
    parser = argparse.ArgumentParser(description="Build prompt payloads from trawled jobs.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input trawl results xlsx")
    parser.add_argument("--resume", type=Path, default=DEFAULT_RESUME_PATH, help="Base resume text file")
    parser.add_argument("--output", type=Path, default=None, help="Output jsonl path")
    args = parser.parse_args()

    output_path = args.output or _default_output_path()

    written, skipped = build_prompt_records(
        input_path=args.input,
        resume_path=args.resume,
        output_path=output_path,
    )
    print(f"Prompt pipeline complete: written={written}, skipped={skipped}, output={output_path}")


if __name__ == "__main__":
    main()
