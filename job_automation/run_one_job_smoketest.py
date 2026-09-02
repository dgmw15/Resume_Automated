"""
run_one_job_smoketest.py — Scoped, single-job end-to-end pipeline test.

Runs the REAL pipeline for exactly ONE job listing instead of main.py's
indefinite run_forever() loop: login -> scrape one listing -> fetch its JD
-> employment filter -> JD validation -> AI tailoring -> keyword coverage
-> ATS critic -> DOCX render. Everything is written to the real
Database.xlsx via the real ExcelTracker, so results show up in
web_ui/app.py exactly like a normal pipeline run would.

MUST be run locally (needs a real browser window for SingPass + real
network access) — this cannot run in a cloud sandbox.

Usage:
    .venv\\Scripts\\activate
    python run_one_job_smoketest.py                    # role "Data Analyst"
    python run_one_job_smoketest.py --role "Data Engineer"
    python run_one_job_smoketest.py --skip-login        # reuse saved session, no login prompt

What it does NOT do: touch run_forever()'s scrape/validate/batch phases for
every configured role, or process more than one row through the AI tailor
even if other rows are already queued in Database.xlsx.
"""
from __future__ import annotations

import argparse
import asyncio
import logging
import sys
from pathlib import Path

import openpyxl
import yaml

from adapters.careersfuture import CareersFutureAdapter
from ai.critic import AtsCritic
from ai.employment_filter import EmploymentFilter
from ai.jd_validator import build_validator_from_config
from ai.provider_router import ProviderRouter
from ai.tailor import ResumeTailor
from core.batch_processor import BatchProcessor, BASE_RESUME_PATH
from core.session_manager import SessionManager
from data.models import JobStatus
from data.tracker import DB_PATH, ExcelTracker
from output.docx_renderer import build_renderer_from_config

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("smoketest")

CONFIG_PATH = Path("config.yaml")


def _already_scraped_urls() -> set[str]:
    """URLs already present in Database.xlsx, so a re-run picks the NEXT listing
    instead of re-processing the same one (e.g. after a validation failure)."""
    if not DB_PATH.exists():
        return set()
    wb = openpyxl.load_workbook(DB_PATH, read_only=True)
    ws = wb["Jobs"]
    col = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}
    url_idx = col["url"] - 1
    return {
        row[url_idx]
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[url_idx]
    }


def _first_unseen(listings, seen_urls: set[str]):
    for listing in listings:
        if listing.url not in seen_urls:
            return listing
    return None


async def main(role: str, skip_login: bool) -> int:
    if not BASE_RESUME_PATH.exists():
        logger.error("base_resume.txt not found at repo root — nothing to tailor with.")
        return 1

    config = yaml.safe_load(CONFIG_PATH.read_text())
    tracker = ExcelTracker()
    session_manager = SessionManager()
    await session_manager.start()

    try:
        ctx = await session_manager.get_context("careersfuture")
        credentials = config.get("credentials", {}).get("careersfuture", {})
        adapter = CareersFutureAdapter(ctx, credentials)

        if not skip_login:
            logger.info("Opening CareersFuture for login — complete SingPass in the browser window.")
            await adapter.login()
            await session_manager.save_state("careersfuture")
            logger.info("Login complete, session saved.")
        else:
            logger.info("--skip-login passed: reusing saved session, no login prompt.")

        logger.info("Scraping page 1 for role=%r ...", role)
        listings = await adapter.scrape_page(role, page_num=1)
        if not listings:
            logger.error("No listings found for role=%r. Try a different --role.", role)
            return 1

        job = _first_unseen(listings, _already_scraped_urls())
        if job is None:
            logger.error(
                "All %d listings on page 1 are already in Database.xlsx. "
                "Delete a row, use a different --role, or clear old test rows.",
                len(listings),
            )
            return 1

        logger.info("Selected job: %r at %r (%s)", job.role, job.company, job.url)
        job.raw_description = await adapter.get_job_description(job.url)
        job.status = JobStatus.SCRAPED
        tracker.append(job)

        # --- Employment filter (same gate the real pipeline uses) ---
        ef_cfg = config.get("employment_filter", {})
        if ef_cfg:
            emp_filter = EmploymentFilter.from_config(config)
            result = emp_filter.classify(title=job.role, description=job.raw_description)
            tracker.mark_employment_filter(
                job_id=job.id, status=result.status, reason=result.reason,
                emp_type_raw=result.employment_type_raw,
                emp_type_normalized=result.employment_type_normalized,
            )
            if result.status == "FILTERED":
                logger.warning("Job FILTERED by employment_filter: %s — stopping here.", result.reason)
                logger.warning("Re-run the script to try the next unseen listing.")
                return 0

        # --- JD validation (same deterministic keyword gate) ---
        default_role = config.get("ai", {}).get("default_role", "analyst")
        validator = build_validator_from_config(config, role=default_role)
        vresult = validator.validate(job.raw_description)
        if not vresult.is_pass:
            tracker.mark_validation_result(
                job.id, status=JobStatus.VALIDATION_FAILED_NON_TECH,
                score=vresult.score, reason=vresult.reason,
            )
            logger.warning("Job FAILED JD validation (score=%d): %s", vresult.score, vresult.reason)
            logger.warning("Re-run the script to try the next unseen listing.")
            return 0

        tracker.mark_validation_result(
            job.id, status=JobStatus.VALIDATION_PASSED, score=vresult.score, reason=vresult.reason,
        )
        logger.info("Job PASSED validation (score=%d). Queuing for AI tailoring...", vresult.score)
        tracker.mark_batch_queued(job.id)

        # --- AI tailoring + coverage + critic + DOCX, scoped to this one row ---
        router = ProviderRouter(config)
        tailor = ResumeTailor(router)
        critic = AtsCritic(router) if config.get("resume_tailoring", {}).get("enable_ats_critic", True) else None
        docx_renderer = build_renderer_from_config(config)
        batch_processor = BatchProcessor(
            tracker, tailor, config.get("batch", {}), docx_renderer=docx_renderer, critic=critic,
        )

        base_resume = BASE_RESUME_PATH.read_text(encoding="utf-8")
        row = tracker.list_rows_by_status(JobStatus.BATCH_QUEUED)
        row = next((r for r in row if r.get("id") == job.id), None)
        if row is None:
            logger.error("Could not re-read the queued row from Database.xlsx — aborting.")
            return 1

        # Calling _process_row directly (rather than run_once()) guarantees exactly
        # this one job is processed, even if other rows are sitting BATCH_QUEUED
        # from an earlier run.
        success = await batch_processor._process_row(row, base_resume)

        if not success:
            logger.error("AI tailoring failed for this job — check the log above.")
            return 1

        final_row = next(
            r for r in tracker.list_rows_by_status(JobStatus.DOCX_READY)
            + tracker.list_rows_by_status(JobStatus.TAILORED_TEXT_READY)
            + tracker.list_rows_by_status(JobStatus.DOCX_GENERATION_FAILED)
            if r.get("id") == job.id
        )
        print()
        print("=" * 60)
        print("SMOKE TEST RESULT")
        print("=" * 60)
        print(f"  Job:              {final_row.get('role')} @ {final_row.get('company')}")
        print(f"  Status:           {final_row.get('status')}")
        print(f"  AI provider used: {final_row.get('ai_provider_used')}")
        print(f"  Cost (USD):       {final_row.get('cost_usd')}")
        print(f"  Keyword coverage: {final_row.get('keyword_coverage_score')}%")
        print(f"  Missing keywords: {final_row.get('keyword_coverage_missing')}")
        print(f"  ATS verdict:      {final_row.get('ats_verdict')}")
        print(f"  DOCX path:        {final_row.get('docx_path')}")
        print()
        print("Open web_ui/app.py (python web_ui/app.py -> http://localhost:8765) to review it.")
        return 0

    finally:
        await session_manager.stop()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--role", default="Data Analyst", help="Search role/keyword (default: 'Data Analyst')")
    parser.add_argument("--skip-login", action="store_true", help="Reuse saved session, skip the login prompt")
    args = parser.parse_args()
    sys.exit(asyncio.run(main(args.role, args.skip_login)))
